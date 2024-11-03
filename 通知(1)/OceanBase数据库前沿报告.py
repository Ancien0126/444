import requests
from bs4 import BeautifulSoup
import re
import os
from openpyxl import Workbook, load_workbook
from openpyxl.styles import Alignment


file_path = '通知.xlsx'

# 检查文件是否存在
if os.path.exists(file_path):
    wb = load_workbook(file_path)
    ws = wb.active
else:
    wb = Workbook()
    ws = wb.active
    excle_headers = ['标题', '时间', '内容', '链接']
    ws.append(excle_headers)
    wb.save(file_path)
for cell in ws['1:1']:
    cell.alignment = Alignment(horizontal='center', vertical='center')


url = "https://www.cs.sdu.edu.cn/info/1055/6396.htm"
response = requests.get(url)

html_text = response.text

# 将源码编码后的字符串解码为原始字符串
html_text = html_text.encode('latin1').decode('utf-8')
# print(html_text)


soup = BeautifulSoup(html_text, 'html.parser')

# 提取通知标题,“find_all('h2')”--找到所有的h2标签，[1]--取列表里的第二个元素
title = soup.find_all('h2')[1].text.strip()
print("活动标题：", title)

# 提取发布时间
time_span = soup.find_all('span')
datetime = None
for ti in time_span:
    if "发布时间" in ti.get_text(): 
        date = ti.get_text()
        # 使用正则表达式匹配日期格式/找到匹配项返回该匹配的字符串
        datetime = re.search(r"\d{4}-\d{2}-\d{2}", date).group()

if datetime:
    print("发布时间：", datetime)


# 提取通知内容:找到所有属性为"style"，其对应属性值为"color: rgb(63, 63, 63);"的标签
content_paragraphs = soup.find_all(attrs={"style": "color: rgb(63, 63, 63);"})

content = ''.join(i.get_text() for i in content_paragraphs)
print("详细内容：", content)


ws.append([title, datetime, content, url])

wb.save(file_path)
print(">>>内容已写入表格，文件已保存")

